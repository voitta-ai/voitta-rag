"""Parser for Microsoft Excel documents (.xlsx and .xls)."""

from pathlib import Path

from .base import BaseParser, ParserResult

# Maximum rows to extract per sheet
MAX_ROWS_PER_SHEET = 100


class XlsxParser(BaseParser):
    """Parser for Excel files (.xlsx and .xls)."""

    extensions = [".xlsx", ".xls"]

    def parse(self, file_path: Path) -> ParserResult:
        """Parse an Excel workbook to Markdown (all sheets, max 100 rows each)."""
        ext = file_path.suffix.lower()

        if ext == ".xlsx":
            return self._parse_xlsx(file_path)
        elif ext == ".xls":
            return self._parse_xls(file_path)
        else:
            return ParserResult.failure(f"Unsupported Excel format: {ext}")

    def _parse_xlsx(self, file_path: Path) -> ParserResult:
        """Parse .xlsx files using openpyxl."""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError:
            return ParserResult.failure("openpyxl not installed. Run: pip install openpyxl")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except InvalidFileException:
            return ParserResult.failure(f"Invalid or corrupted XLSX file: {file_path}")
        except Exception as e:
            return ParserResult.failure(f"Failed to open XLSX file: {e}")

        lines = []
        metadata = {}

        # Extract document properties
        try:
            props = wb.properties
            if props.title:
                metadata["title"] = props.title
            if props.creator:
                metadata["author"] = props.creator
        except Exception:
            pass

        # Add workbook title
        if metadata.get("title"):
            lines.append(f"# {metadata['title']}")
        else:
            lines.append(f"# {file_path.stem}")
        lines.append("")

        # Process all sheets
        sheet_names = wb.sheetnames
        metadata["sheets"] = sheet_names

        for sheet_name in sheet_names:
            sheet = wb[sheet_name]

            # Clear sheet header with name for chunking context
            lines.append(f"## Sheet: {sheet_name}")
            lines.append("")

            # Get rows with data (limited to MAX_ROWS_PER_SHEET)
            rows_data = []
            max_col = 0
            row_count = 0

            for row in sheet.iter_rows():
                if row_count >= MAX_ROWS_PER_SHEET:
                    break

                row_values = []
                has_content = False
                for cell in row:
                    value = cell.value
                    if value is not None:
                        has_content = True
                        # Convert to string and clean
                        str_value = str(value).strip().replace("\n", " ").replace("|", "\\|")
                        row_values.append(str_value)
                    else:
                        row_values.append("")

                if has_content:
                    rows_data.append(row_values)
                    max_col = max(max_col, len(row_values))
                    row_count += 1

            if not rows_data:
                lines.append("*(Empty sheet)*")
                lines.append("")
                continue

            # Normalize all rows to same column count
            for row in rows_data:
                while len(row) < max_col:
                    row.append("")

            # Build markdown table
            # Header row (first row)
            header = rows_data[0] if rows_data else []
            lines.append("| " + " | ".join(header) + " |")

            # Separator
            lines.append("| " + " | ".join(["---"] * max_col) + " |")

            # Data rows
            for row in rows_data[1:]:
                lines.append("| " + " | ".join(row) + " |")

            # Note if truncated
            if row_count >= MAX_ROWS_PER_SHEET:
                lines.append("")
                lines.append(f"*(Showing first {MAX_ROWS_PER_SHEET} rows)*")

            lines.append("")

        wb.close()

        content = "\n".join(lines).strip()
        return ParserResult(content=content, metadata=metadata)

    def _parse_xls(self, file_path: Path) -> ParserResult:
        """Parse .xls files using xlrd."""
        try:
            import xlrd
        except ImportError:
            return ParserResult.failure("xlrd not installed. Run: pip install xlrd")

        try:
            wb = xlrd.open_workbook(file_path)
        except xlrd.XLRDError as e:
            return ParserResult.failure(f"Invalid or corrupted XLS file: {e}")
        except Exception as e:
            return ParserResult.failure(f"Failed to open XLS file: {e}")

        lines = []
        metadata = {}

        # Add workbook title (xls doesn't have properties like xlsx)
        lines.append(f"# {file_path.stem}")
        lines.append("")

        # Process all sheets
        sheet_names = wb.sheet_names()
        metadata["sheets"] = sheet_names

        for sheet_name in sheet_names:
            sheet = wb.sheet_by_name(sheet_name)

            # Clear sheet header with name for chunking context
            lines.append(f"## Sheet: {sheet_name}")
            lines.append("")

            # Get dimensions
            nrows = min(sheet.nrows, MAX_ROWS_PER_SHEET)
            ncols = sheet.ncols

            if nrows == 0 or ncols == 0:
                lines.append("*(Empty sheet)*")
                lines.append("")
                continue

            # Get rows with data
            rows_data = []
            for row_idx in range(nrows):
                row_values = []
                has_content = False
                for col_idx in range(ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if value is not None and value != "":
                        has_content = True
                        # Convert to string and clean
                        str_value = str(value).strip().replace("\n", " ").replace("|", "\\|")
                        row_values.append(str_value)
                    else:
                        row_values.append("")

                if has_content:
                    rows_data.append(row_values)

            if not rows_data:
                lines.append("*(Empty sheet)*")
                lines.append("")
                continue

            # Find max columns with actual content
            max_col = max(len(row) for row in rows_data)

            # Normalize all rows to same column count
            for row in rows_data:
                while len(row) < max_col:
                    row.append("")

            # Build markdown table
            # Header row (first row)
            header = rows_data[0] if rows_data else []
            lines.append("| " + " | ".join(header) + " |")

            # Separator
            lines.append("| " + " | ".join(["---"] * max_col) + " |")

            # Data rows
            for row in rows_data[1:]:
                lines.append("| " + " | ".join(row) + " |")

            # Note if truncated
            if sheet.nrows > MAX_ROWS_PER_SHEET:
                lines.append("")
                lines.append(f"*(Showing first {MAX_ROWS_PER_SHEET} rows)*")

            lines.append("")

        content = "\n".join(lines).strip()
        return ParserResult(content=content, metadata=metadata)
